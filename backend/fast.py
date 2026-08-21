import os
from pathlib import Path
from typing import List, Optional
from dotenv import load_dotenv
from fastapi import FastAPI, Query, UploadFile, File, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import psycopg2
import psycopg2.errors
from psycopg2.extras import RealDictCursor
import pandas as pd
from sqlalchemy import create_engine, text
from sqlalchemy.engine import URL
import io
import re
from collections import defaultdict
from datetime import date, datetime
import openpyxl

from auth import (
    hash_password,
    verify_password,
    create_access_token,
    get_current_user,
    require_roles,
)

app = FastAPI(title="Network Performance API (Enterprise Version)")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==========================================
# 1. Database Configuration (Supabase / Postgres via .env)
# ==========================================
# โหลดค่าจากไฟล์ .env โดยชี้ path ตรงไปที่โฟลเดอร์เดียวกับไฟล์นี้เสมอ
# (ไม่พึ่ง current working directory เพราะบางทีรัน uvicorn จากคนละโฟลเดอร์ ทำให้หา .env ไม่เจอ)
# ถ้ายังไม่มีไฟล์ .env: copy .env.example -> .env แล้วกรอกรหัสผ่าน Supabase จริง
ENV_PATH = Path(__file__).resolve().parent / ".env"
# override=True: บังคับให้ค่าใน .env ทับ environment variable เดิมที่อาจค้างอยู่ในระบบ/เทอร์มินัล
# (ถ้าไม่ใส่ override, load_dotenv จะไม่ทับค่าที่มี env var ชื่อเดียวกันตั้งอยู่แล้วในเครื่อง
#  ซึ่งเป็นสาเหตุคลาสสิกที่ทำให้ .env ดูเหมือน "ไม่ทำงาน" ทั้งที่ไฟล์ถูกต้องแล้ว)
load_dotenv(dotenv_path=ENV_PATH, override=True)

def _env(key, default=None):
    """ อ่านค่า env var โดยถือว่าค่าว่าง ('' หรือช่องว่างล้วน) เหมือนกับไม่มีค่า -> ใช้ default แทน """
    val = os.getenv(key)
    if val is None or val.strip() == "":
        return default
    return val.strip()

DB_USER = _env("DB_USER", "postgres")
DB_PASSWORD = _env("DB_PASSWORD")
DB_HOST = _env("DB_HOST")
DB_PORT = _env("DB_PORT", "5432")
DB_NAME = _env("DB_NAME", "postgres")
DB_SSLMODE = _env("DB_SSLMODE", "require")  # Supabase ต้องการ SSL เสมอ

if not DB_PASSWORD or not DB_HOST:
    raise RuntimeError(
        f"ไม่พบ DB_PASSWORD หรือ DB_HOST ที่ถูกต้องใน {ENV_PATH}\n"
        "ตรวจสอบว่า:\n"
        "  1) มีไฟล์ .env อยู่ในโฟลเดอร์เดียวกับ fast.py จริง (ไม่ใช่ .env.example)\n"
        "  2) แต่ละบรรทัดเป็นรูปแบบ KEY=value ไม่มีช่องว่างรอบเครื่องหมาย = และไม่มีเครื่องหมายคำพูดครอบ\n"
        "  3) DB_PORT ต้องเป็นตัวเลข เช่น 5432 (ห้ามเว้นว่าง)"
    )

try:
    int(DB_PORT)
except (TypeError, ValueError):
    raise RuntimeError(
        f"DB_PORT ใน {ENV_PATH} ไม่ใช่ตัวเลขที่ถูกต้อง (ค่าที่อ่านได้: {DB_PORT!r}) "
        "ปกติควรเป็น 5432 — ตรวจสอบว่าไม่มีช่องว่างหรือพิมพ์ตกหล่นในไฟล์ .env"
    )

# สร้าง URL ด้วย URL.create() แทนการต่อ f-string ตรงๆ เพราะ DB_PASSWORD ที่ Supabase
# generate ให้มักมีอักขระพิเศษ (เช่น @ : / # ?) ซึ่งถ้าต่อ string เองด้วยมือจะทำให้ parser
# อ่านตำแหน่ง host/port ผิดเพี้ยน (อาการที่เจอ: port กลายเป็นค่าว่าง) URL.create() จะ escape
# อักขระพิเศษเหล่านี้ให้ถูกต้องอัตโนมัติ
db_url = URL.create(
    drivername="postgresql",
    username=DB_USER,
    password=DB_PASSWORD,
    host=DB_HOST,
    port=int(DB_PORT),
    database=DB_NAME,
    query={"sslmode": DB_SSLMODE},
)

print(f"[db] เชื่อมต่อไปที่ {DB_USER}@{DB_HOST}:{DB_PORT}/{DB_NAME} (sslmode={DB_SSLMODE})")
print(f"[db] DB_PASSWORD ที่อ่านได้ยาว {len(DB_PASSWORD)} ตัวอักษร (ไม่โชว์ค่าจริงเพื่อความปลอดภัย)")

try:
    engine = create_engine(db_url)
except Exception as e:
    raise RuntimeError(
        f"สร้าง database engine ไม่สำเร็จ: {e}\n"
        f"ตรวจสอบค่าใน {ENV_PATH} อีกครั้ง (DB_HOST={DB_HOST!r}, DB_PORT={DB_PORT!r}, DB_NAME={DB_NAME!r})"
    ) from e

def get_db_connection():
    return psycopg2.connect(
        user=DB_USER, password=DB_PASSWORD, host=DB_HOST, port=int(DB_PORT),
        database=DB_NAME, sslmode=DB_SSLMODE
    )

# ==========================================
# 1b. Pydantic models (auth / user management / export)
# ==========================================
class LoginRequest(BaseModel):
    username: str
    password: str

class CreateUserRequest(BaseModel):
    username: str
    password: str
    role: str  # 'admin' หรือ 'visitor' เท่านั้น (owner มอบสิทธิ์ผ่าน endpoint แยก)

class ChangeRoleRequest(BaseModel):
    new_role: str  # 'owner' | 'admin' | 'visitor'
    owner_password: str  # รหัสผ่านของ owner ที่ล็อกอินอยู่ ใช้ยืนยันตัวตนก่อนเปลี่ยนสิทธิ์

class ResetPasswordRequest(BaseModel):
    new_password: str

# ==========================================
# 2. Base Configuration & Table Registry
# ==========================================
# ส่วนกลางสำหรับคอลัมน์ที่มีในทุกๆ ไฟล์ Performance
BASE_PERF_MAPPING = {
    'Begin Time': 'begin_time', 'End Time': 'end_time', 'Granularity': 'granularity',
    'ME': 'me_name', 'ME IP': 'me_ip', 'Measure Object': 'measure_object'
}

# Registry ตรวจสอบไฟล์อัตโนมัติ (ชื่อตารางอ้างอิงตามฐานข้อมูลจริง)
TABLE_REGISTRY = {
    "alarm_history": {
        "table_name": "fm-history",
        "file_keyword": "fm-history",
        "required_columns": ["Alarm ID", "Occurrence Time", "Alarm Code"],
        "conflict_target": "alarm_id",
        "is_performance": False,
        "mapping": {
            'Alarm ID': 'alarm_id', 'ME': 'me_name', 'ME IP': 'me_ip',
            'Alarm Code': 'alarm_code', 'Alarm Code Name': 'alarm_code_name',
            'Alarm Severity': 'alarm_severity', 'Occurrence Time': 'occurrence_time',
            'Clear Time': 'clear_time', 'Duration Time': 'duration_time'
        }
    },
    "cpu_perf": {
        "table_name": "cpu_performance",
        "file_keyword": "CPU ratio",
        "required_columns": ["Max CPU utilization ratio", "CPU utilization ratio", "RAM utilization ratio"],
        "conflict_target": "begin_time, me_ip, measure_object",
        "is_performance": True,
        "mapping": {**BASE_PERF_MAPPING, **{
            'Max CPU utilization ratio': 'cpu_max_ratio', 'Min CPU utilization ratio': 'cpu_min_ratio',
            'CPU utilization ratio': 'cpu_avg_ratio', 'RAM utilization ratio': 'ram_avg_ratio',
            'Max RAM utilization ratio': 'ram_max_ratio', 'Min RAM utilization ratio': 'ram_min_ratio'
        }}
    },
    "client_card_perf": {
        "table_name": "client-card_performance",
        "file_keyword": "Client card performance",
        "required_columns": ["Max Value of Output Optical Power(dBm)", "Output Optical Power (dBm)"],
        "conflict_target": "begin_time, me_ip, measure_object",
        "is_performance": True,
        "mapping": {**BASE_PERF_MAPPING, **{
            'Max Value of Output Optical Power(dBm)': 'max_out_power', 'Min Value of Output Optical Power(dBm)': 'min_out_power',
            'Input Optical Power(dBm)': 'input_power', 'Output Optical Power (dBm)': 'output_power'
        }}
    },
    "fan_perf": {
        "table_name": "fan_performance",
        "file_keyword": "FAN ratio",
        "required_columns": ["Value of Fan Rotate Speed(Rps)"],
        "conflict_target": "begin_time, me_ip, measure_object",
        "is_performance": True,
        "mapping": {**BASE_PERF_MAPPING, **{
            'Max Value of Fan Rotate Speed(Rps)': 'max_fan_speed', 'Min Value of Fan Rotate Speed(Rps)': 'min_fan_speed',
            'Value of Fan Rotate Speed(Rps)': 'fan_speed'
        }}
    },
    "line_card_perf": {
        "table_name": "line-card_performance",
        "file_keyword": "Line cards performance",
        "required_columns": ["Instant BER After FEC", "Instant BER Before FEC"],
        "conflict_target": "begin_time, me_ip, measure_object",
        "is_performance": True,
        "mapping": {**BASE_PERF_MAPPING, **{
            'Instant BER After FEC': 'ber_after_fec', 'Instant BER Before FEC': 'ber_before_fec',
            'Input Optical Power(dBm)': 'input_power', 'Output Optical Power (dBm)': 'output_power'
        }}
    },
    "msu_perf": {
        "table_name": "msu_performance",
        "file_keyword": "MSU performance",
        "required_columns": ["Laser Bias Current(mA)"],
        "conflict_target": "begin_time, me_ip, measure_object",
        "is_performance": True,
        "mapping": {**BASE_PERF_MAPPING, **{
            'Max Value of Laser Bias Current(mA)': 'max_laser_current', 'Min Value of Laser Bias Current(mA)': 'min_laser_current', 
            'Laser Bias Current(mA)': 'laser_current'
        }}
    },
    "osc_power": {
        "table_name": "osc_performance",
        "file_keyword": "OSC Power Flapping",
        "required_columns": ["Max Value of Input Optical Power(dBm)", "Input Optical Power(dBm)"],
        "conflict_target": "begin_time, me_ip, measure_object",
        "is_performance": True,
        "mapping": {**BASE_PERF_MAPPING, **{
            'Max Value of Input Optical Power(dBm)': 'max_input_power', 'Min Value of Input Optical Power(dBm)': 'min_input_power',
            'Input Optical Power(dBm)': 'input_power'
        }}
    }
}

# ==========================================
# 2b. Threshold sheet registry (ไฟล์เดียวกัน แต่คนละ sheet)
#     key = ชื่อ sheet จริงในไฟล์ Excel, value = prefix กัน metric_key ชนกัน
# ==========================================
THRESHOLD_SHEETS = {
    "Control Ratio": "cpu",   # sheet นี้คือ threshold ของ CPU (ยืนยันจากผู้ใช้)
    "Fan Ratio": "fan",
    "MSU": "msu",
    "Line board": "line",
    "Client board": "client",
}

THRESHOLD_IDENTIFIER_HEADERS = {
    "site name": "site_name",
    "me": "me_name",
    "me ip": "me_ip",
    "measure object": "measure_object",
}
THRESHOLD_MAX_LABELS = {"maximum threshold", "max threshold", "threshold"}
THRESHOLD_MIN_LABELS = {"minimum threshold", "minimun threshold", "min threshold"}  # 'minimun' = typo ที่พบจริงในไฟล์

# ==========================================
# 3. Helper Functions
# ==========================================
def extract_measure_object_details(df: pd.DataFrame) -> pd.DataFrame:
    """ ฟังก์ชันสกัดข้อมูลบอร์ดและพอร์ตจากคอลัมน์ measure_object """
    if 'measure_object' in df.columns:
        df['board_name'] = df['measure_object'].str.extract(r'^([^\[]+)')
        df['location_path'] = df['measure_object'].str.extract(r'\[([0-9-]+)\]')
        df['sub_function'] = df['measure_object'].str.extract(r'\]-([^:(]+)')
    return df

def execute_pg_upsert(df: pd.DataFrame, table_name: str, conflict_target: str):
    """ ฟังก์ชันบันทึกข้อมูลเข้าฐานข้อมูลแบบกันข้อมูลซ้ำ (Upsert) """
    staging_table = f"temp_staging_{table_name.replace('-', '_')}"
    
    # เพิ่ม " รอบชื่อคอลัมน์ เพื่อรองรับอักษรพิมพ์เล็ก-ใหญ่ใน PostgreSQL อย่างถูกต้อง
    columns_str = ", ".join([f'"{c}"' for c in df.columns])
    
    with engine.begin() as conn:
        # 1. นำข้อมูลไปพักที่ Staging Table
        df.to_sql(name=staging_table, con=conn, if_exists='replace', index=False, method='multi', chunksize=5000)
        
        # 2. Insert เข้าตารางจริง หากซ้ำให้ข้าม (DO NOTHING)
        upsert_sql = f"""
            INSERT INTO "{table_name}" ({columns_str})
            SELECT {columns_str} FROM "{staging_table}"
            ON CONFLICT ({conflict_target}) 
            DO NOTHING;
        """
        conn.execute(text(upsert_sql))
        
        # 3. ลบตารางพักข้อมูลทิ้ง
        conn.execute(text(f'DROP TABLE IF EXISTS "{staging_table}";'))


# ==========================================
# 3b. Threshold parsing helpers
#     (อ่าน sheet ที่มีคอลัมน์ 'Maximum threshold' / 'Minimum threshold' (หรือพิมพ์ผิดเป็น
#      'Minimun threshold' ตามที่พบจริงในไฟล์) สลับกับคอลัมน์ค่าจริง วนซ้ำหลายรอบ
#      บาง sheet เช่น MSU มีแค่ Maximum threshold อย่างเดียว ไม่มี Minimum เลย -- รองรับ optional)
# ==========================================
def _norm_header(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _metric_key_from_label(label: str) -> str:
    base = re.sub(r"\(.*?\)", "", label)
    base = re.sub(r"[^a-zA-Z0-9]+", "_", base).strip("_").lower()
    return base or "metric"


def _extract_unit(label: str):
    m = re.search(r"\(([^)]+)\)", label)
    return m.group(1) if m else None


def parse_threshold_sheet(ws, sheet_name, prefix, header_row_idx=2):
    """
    อ่าน 1 sheet threshold คืนค่า list[dict] (ยังไม่ dedupe ข้ามวันที่)

    หมายเหตุ perf: อ่านทั้ง sheet ด้วย iter_rows() ครั้งเดียว (sequential) แทนการสุ่มเข้าถึง
    ทีละ cell ด้วย ws.cell(row=,column=) เพราะ workbook ที่ต้องเปิดแบบ read_only=True (จำเป็นสำหรับ
    ไฟล์ใหญ่ที่มี sheet อื่นในไฟล์เดียวกันซึ่งหนักมาก เช่น sheet ที่มีเป็นหมื่นแถว) จะช้ามากถ้าสุ่มเข้าถึงทีละ cell
    """
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < header_row_idx:
        return []

    header_row = all_rows[header_row_idx - 1]
    max_col = len(header_row)
    max_row = len(all_rows)
    headers = [_norm_header(v) for v in header_row]

    context_row = all_rows[header_row_idx - 2] if header_row_idx - 2 >= 0 else None

    def date_context_for_col(col_idx):
        if not context_row:
            return None
        c = col_idx - 1
        while c >= 0:
            val = context_row[c] if c < len(context_row) else None
            if val not in (None, ""):
                if isinstance(val, (datetime, date)):
                    return val if isinstance(val, date) else val.date()
                return val
            c -= 1
        return None

    id_cols, first_threshold_col = {}, None
    for idx, h in enumerate(headers, start=1):
        low = h.lower()
        if low in THRESHOLD_MAX_LABELS:
            first_threshold_col = idx
            break
        if low in THRESHOLD_IDENTIFIER_HEADERS:
            id_cols[idx] = THRESHOLD_IDENTIFIER_HEADERS[low]

    if first_threshold_col is None:
        return []

    blocks = []
    c = first_threshold_col
    while c <= max_col:
        h = headers[c - 1].lower()
        if h in THRESHOLD_MAX_LABELS:
            max_c = c
            next_h = headers[c].lower() if c + 1 <= max_col else ""
            if next_h in THRESHOLD_MIN_LABELS:
                min_c, val_c = c + 1, c + 2
            else:
                min_c, val_c = None, c + 1

            metric_label = headers[val_c - 1] if val_c <= max_col else ""
            if not metric_label or metric_label.lower() in THRESHOLD_MAX_LABELS or metric_label.lower() in THRESHOLD_MIN_LABELS:
                metric_label = f"metric_{len(blocks) + 1}"
                next_c = (min_c or max_c) + 1
            else:
                next_c = val_c + 1

            blocks.append({
                "max_col": max_c, "min_col": min_c,
                "metric_label": metric_label,
                "date": date_context_for_col(max_c),
            })
            c = next_c
        else:
            c += 1

    if not blocks:
        return []

    records = []
    for r_idx in range(header_row_idx, max_row):
        row = all_rows[r_idx]
        row_vals = {field: (row[col_idx - 1] if col_idx - 1 < len(row) else None)
                    for col_idx, field in id_cols.items()}
        measure_object, me_name = row_vals.get("measure_object"), row_vals.get("me_name")
        if not measure_object or not me_name:
            continue
        for b in blocks:
            max_v = row[b["max_col"] - 1] if b["max_col"] - 1 < len(row) else None
            min_v = (row[b["min_col"] - 1] if (b["min_col"] and b["min_col"] - 1 < len(row)) else None)
            if max_v is None and min_v is None:
                continue
            records.append({
                "source_sheet": sheet_name,
                "metric_key": f"{prefix}_{_metric_key_from_label(b['metric_label'])}",
                "metric_label": b["metric_label"],
                "unit": _extract_unit(b["metric_label"]),
                "site_name": row_vals.get("site_name"),
                "me_name": me_name,
                "me_ip": row_vals.get("me_ip"),
                "measure_object": measure_object,
                "min_threshold": min_v,
                "max_threshold": max_v,
                "source_date": b["date"],
            })
    return records


def merge_threshold_records(records):
    """ ยุบหลายวันที่ให้เหลือ 1 ค่าต่อ (metric_key, me_name, measure_object); ใช้ค่าล่าสุดถ้าไม่ตรงกัน """
    groups = defaultdict(list)
    for rec in records:
        groups[(rec["metric_key"], rec["me_name"], rec["measure_object"])].append(rec)

    merged = []
    for _, recs in groups.items():
        distinct = {(r["min_threshold"], r["max_threshold"]) for r in recs}
        if len(distinct) > 1:
            recs_sorted = sorted(recs, key=lambda r: (r["source_date"] is None, r["source_date"]))
            chosen = recs_sorted[-1]
        else:
            chosen = recs[0]
        mo = chosen["measure_object"]
        board_name = re.match(r"^([^\[]+)", mo)
        location_path = re.search(r"\[([0-9-]+)\]", mo)
        sub_function = re.search(r"\]-([^:()]+)", mo)
        merged.append({
            **chosen,
            "board_name": board_name.group(1).strip() if board_name else None,
            "location_path": location_path.group(1) if location_path else None,
            "sub_function": sub_function.group(1).strip() if sub_function else None,
        })
    return merged


def upsert_thresholds_df(df: pd.DataFrame):
    if df.empty:
        return
    cols = ["source_sheet", "metric_key", "metric_label", "unit", "site_name",
            "me_name", "me_ip", "measure_object", "board_name", "location_path",
            "sub_function", "min_threshold", "max_threshold", "source_date"]
    df = df[cols]
    staging = "temp_staging_thresholds"
    columns_str = ", ".join(f'"{c}"' for c in cols)

    with engine.begin() as conn:
        df.to_sql(staging, con=conn, if_exists="replace", index=False, method="multi", chunksize=2000)
        conn.execute(text(f"""
            INSERT INTO "thresholds" ({columns_str})
            SELECT {columns_str} FROM "{staging}"
            ON CONFLICT (metric_key, me_name, measure_object)
            DO UPDATE SET
                min_threshold = EXCLUDED.min_threshold,
                max_threshold = EXCLUDED.max_threshold,
                unit          = EXCLUDED.unit,
                metric_label  = EXCLUDED.metric_label,
                source_sheet  = EXCLUDED.source_sheet,
                board_name    = EXCLUDED.board_name,
                location_path = EXCLUDED.location_path,
                sub_function  = EXCLUDED.sub_function,
                source_date   = EXCLUDED.source_date,
                updated_at    = now();
        """))
        conn.execute(text(f'DROP TABLE IF EXISTS "{staging}";'))


# ==========================================
# 4. Authentication Endpoints
# ==========================================
@app.post("/api/auth/login")
def login(body: LoginRequest):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute(
            "SELECT id, username, password_hash, role FROM users WHERE username = %s",
            (body.username,),
        )
        row = cursor.fetchone()
        cursor.close()
        conn.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"เชื่อมต่อฐานข้อมูลล้มเหลว: {e}")

    if not row or not verify_password(body.password, row["password_hash"]):
        raise HTTPException(status_code=401, detail="ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง")

    token = create_access_token(row["id"], row["username"], row["role"])
    return {
        "status": "success",
        "access_token": token,
        "token_type": "bearer",
        "user": {"id": row["id"], "username": row["username"], "role": row["role"]},
    }


@app.get("/api/auth/me")
def get_me(user: dict = Depends(get_current_user)):
    return {"status": "success", "user": user}


# ==========================================
# 5. User Management Endpoints (Owner เท่านั้น ยกเว้นที่ระบุไว้)
# ==========================================
@app.get("/api/users")
def list_users(user: dict = Depends(require_roles("owner"))):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute(
            "SELECT id, username, role, created_at FROM users ORDER BY "
            "CASE role WHEN 'owner' THEN 0 WHEN 'admin' THEN 1 ELSE 2 END, username"
        )
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return {"status": "success", "data": rows}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.post("/api/users")
def create_user(body: CreateUserRequest, user: dict = Depends(require_roles("owner"))):
    if body.role not in ("admin", "visitor"):
        raise HTTPException(
            status_code=400,
            detail="role ที่สร้างได้ตรงนี้มีแค่ 'admin' หรือ 'visitor' เท่านั้น "
                   "(การมอบสิทธิ์ Owner ทำผ่าน PUT /api/users/{id}/role แยกต่างหาก)",
        )
    if len(body.password) < 8:
        raise HTTPException(status_code=400, detail="รหัสผ่านต้องยาวอย่างน้อย 8 ตัวอักษร")
    if not body.username or len(body.username.strip()) < 3:
        raise HTTPException(status_code=400, detail="ชื่อผู้ใช้ต้องยาวอย่างน้อย 3 ตัวอักษร")

    password_hash = hash_password(body.password)
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO users (username, password_hash, role) VALUES (%s, %s, %s) RETURNING id",
            (body.username.strip(), password_hash, body.role),
        )
        new_id = cursor.fetchone()[0]
        conn.commit()
        cursor.close()
        conn.close()
    except psycopg2.errors.UniqueViolation:
        raise HTTPException(status_code=400, detail=f"มีชื่อผู้ใช้ '{body.username}' อยู่แล้วในระบบ")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    return {"status": "success", "message": f"สร้างผู้ใช้ '{body.username}' (สิทธิ์: {body.role}) สำเร็จ", "id": new_id}


@app.put("/api/users/{target_id}/role")
def change_role(target_id: int, body: ChangeRoleRequest, user: dict = Depends(require_roles("owner"))):
    if body.new_role not in ("owner", "admin", "visitor"):
        raise HTTPException(status_code=400, detail="role ไม่ถูกต้อง ต้องเป็น owner, admin หรือ visitor")

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)

    # ยืนยันตัวตนด้วยรหัสผ่านของ owner ที่ล็อกอินอยู่ ก่อนอนุญาตให้เปลี่ยน role ของใครก็ตาม
    cursor.execute("SELECT password_hash FROM users WHERE id = %s", (user["id"],))
    acting_owner = cursor.fetchone()
    if not acting_owner or not verify_password(body.owner_password, acting_owner["password_hash"]):
        cursor.close()
        conn.close()
        raise HTTPException(status_code=401, detail="รหัสผ่านยืนยันตัวตนไม่ถูกต้อง")

    cursor.execute("SELECT id, username, role FROM users WHERE id = %s", (target_id,))
    target = cursor.fetchone()
    if not target:
        cursor.close()
        conn.close()
        raise HTTPException(status_code=404, detail="ไม่พบผู้ใช้ที่ต้องการแก้ไข")

    try:
        if body.new_role == "owner":
            if target["id"] == user["id"]:
                raise HTTPException(status_code=400, detail="คุณเป็น Owner อยู่แล้ว")
            # โอนสิทธิ์ Owner: ลด role คนเดิมเป็น admin ก่อน แล้วค่อยเลื่อนคนใหม่ขึ้นเป็น owner
            # (ต้องทำ 2 ขั้นตอนนี้ในลำดับนี้เท่านั้น เพราะ DB บังคับว่ามี owner ได้แค่ 1 คน)
            cursor.execute("UPDATE users SET role = 'admin', updated_at = now() WHERE id = %s", (user["id"],))
            cursor.execute("UPDATE users SET role = 'owner', updated_at = now() WHERE id = %s", (target_id,))
            msg = f"โอนสิทธิ์ Owner ให้ '{target['username']}' สำเร็จ (บัญชีของคุณถูกลดเป็น admin โดยอัตโนมัติ)"
        else:
            if target["role"] == "owner":
                raise HTTPException(
                    status_code=400,
                    detail="ต้องโอนสิทธิ์ Owner ให้คนอื่นก่อน (ผ่านการตั้ง role เป็น owner ให้คนนั้น) "
                           "ถึงจะลดสิทธิ์ตัวเองได้ ระบบต้องมี owner อยู่เสมออย่างน้อย 1 คน",
                )
            cursor.execute(
                "UPDATE users SET role = %s, updated_at = now() WHERE id = %s",
                (body.new_role, target_id),
            )
            msg = f"เปลี่ยนสิทธิ์ของ '{target['username']}' เป็น {body.new_role} สำเร็จ"
        conn.commit()
    except HTTPException:
        conn.rollback()
        cursor.close()
        conn.close()
        raise
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))

    cursor.close()
    conn.close()
    return {"status": "success", "message": msg}


@app.post("/api/users/{target_id}/reset-password")
def reset_password(target_id: int, body: ResetPasswordRequest, user: dict = Depends(require_roles("owner"))):
    if len(body.new_password) < 8:
        raise HTTPException(status_code=400, detail="รหัสผ่านต้องยาวอย่างน้อย 8 ตัวอักษร")

    password_hash = hash_password(body.new_password)
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE users SET password_hash = %s, updated_at = now() WHERE id = %s RETURNING username",
        (password_hash, target_id),
    )
    row = cursor.fetchone()
    if not row:
        conn.rollback()
        cursor.close()
        conn.close()
        raise HTTPException(status_code=404, detail="ไม่พบผู้ใช้นี้")
    conn.commit()
    cursor.close()
    conn.close()
    return {"status": "success", "message": f"ตั้งรหัสผ่านใหม่ให้ '{row[0]}' สำเร็จ"}


@app.delete("/api/users/{target_id}")
def delete_user(target_id: int, user: dict = Depends(require_roles("owner"))):
    if target_id == user["id"]:
        raise HTTPException(status_code=400, detail="ไม่สามารถลบบัญชีของตัวเองได้")

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT username, role FROM users WHERE id = %s", (target_id,))
    row = cursor.fetchone()
    if not row:
        cursor.close()
        conn.close()
        raise HTTPException(status_code=404, detail="ไม่พบผู้ใช้นี้")
    if row["role"] == "owner":
        cursor.close()
        conn.close()
        raise HTTPException(status_code=400, detail="ลบบัญชี Owner ไม่ได้ ต้องโอนสิทธิ์ให้คนอื่นก่อน")

    cursor.execute("DELETE FROM users WHERE id = %s", (target_id,))
    conn.commit()
    cursor.close()
    conn.close()
    return {"status": "success", "message": f"ลบผู้ใช้ '{row['username']}' สำเร็จ"}


# ==========================================
# 6. API Endpoints (ข้อมูล Performance เดิม)
# ==========================================
@app.get("/api/metrics")
def get_network_metrics(
    me_ip: str = Query(..., description="ไอพีของอุปกรณ์เครือข่าย"),
    user: dict = Depends(get_current_user),
):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # ดึงจากตาราง cpu_performance เป็นหลัก
        query = """
            SELECT 
                begin_time, me_name, me_ip,
                ROUND((cpu_avg_ratio * 100)::numeric, 2) as cpu_percent,
                ROUND((ram_avg_ratio * 100)::numeric, 2) as ram_percent
            FROM "cpu_performance"
            WHERE me_ip = %s ORDER BY begin_time ;
        """
        cursor.execute(query, (me_ip,))
        data = cursor.fetchall()
        cursor.close()
        conn.close()
        return {"status": "success", "total_records": len(data), "data": data}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/data")
def get_all_data(
    table_name: str = Query("cpu_performance", description="ชื่อตารางที่ต้องการดึงข้อมูล"),
    user: dict = Depends(get_current_user),
):
    """ Endpoint สำหรับดึงข้อมูลจากตารางที่ต้องการไปแสดงหน้าเว็บ (ต้อง login ก่อน ทุก role ดูได้) """
    # ตรวจสอบความปลอดภัย ป้องกัน SQL Injection
    allowed_tables = [config["table_name"] for config in TABLE_REGISTRY.values()]
    if table_name not in allowed_tables:
        return {"status": "error", "message": "ไม่อนุญาตให้เข้าถึงตารางนี้"}

    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # ค้นหาคอลัมน์เวลาของตารางนั้นๆ (เพื่อจัดเรียง)
        order_col = "occurrence_time" if table_name == "fm-history" else "begin_time"
        
        query = f'SELECT * FROM "{table_name}" ORDER BY "{order_col}" '
        cursor.execute(query)
        
        data = cursor.fetchall()
        cursor.close()
        conn.close()
        return {"status": "success", "table": table_name, "data": data}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/api/export")
def export_table_to_excel(
    table_name: str = Query(..., description="ชื่อตารางที่ต้องการ export"),
    user: dict = Depends(require_roles("owner", "admin")),
):
    """ Export ข้อมูลทั้งตารางเป็นไฟล์ .xlsx ให้ดาวน์โหลด (เฉพาะ owner/admin) """
    allowed_tables = [config["table_name"] for config in TABLE_REGISTRY.values()] + ["thresholds"]
    if table_name not in allowed_tables:
        raise HTTPException(status_code=400, detail="ไม่อนุญาตให้ export ตารางนี้")

    try:
        order_col = "occurrence_time" if table_name == "fm-history" else (
            "updated_at" if table_name == "thresholds" else "begin_time"
        )
        df = pd.read_sql(f'SELECT * FROM "{table_name}" ORDER BY "{order_col}"', con=engine)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"ดึงข้อมูลสำหรับ export ล้มเหลว: {e}")

    if df.empty:
        raise HTTPException(status_code=404, detail=f"ตาราง '{table_name}' ไม่มีข้อมูลให้ export")

    # เขียนเป็น .xlsx ในหน่วยความจำ (ไม่สร้างไฟล์ชั่วคราวบนดิสก์)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # ชื่อ sheet ห้ามเกิน 31 ตัวอักษรและห้ามมีอักขระต้องห้ามของ Excel
        safe_sheet_name = re.sub(r"[\\/*?:\[\]]", "_", table_name)[:31]
        df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
    buffer.seek(0)

    export_filename = f"{table_name.replace('-', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{export_filename}"'},
    )


def _process_single_upload(contents: bytes, filename: str) -> dict:
    """
    ประมวลผลไฟล์ .xlsx 1 ไฟล์ (เป็น bytes ที่อ่านมาแล้ว) แล้ว upsert เข้า database ที่ตารางที่ถูกต้อง
    แยกออกมาจาก endpoint เพื่อให้ /api/upload วนลูปเรียกได้หลายไฟล์ในคำขอเดียว
    คืนค่า dict เดียวกับรูปแบบ response เดิมของ endpoint (status/filename/detected_table/...)
    """
    MAX_FILE_SIZE = 50 * 1024 * 1024
    if len(contents) > MAX_FILE_SIZE:
        return {"status": "error", "filename": filename, "message": "ไฟล์มีขนาดใหญ่เกินไป (จำกัดไม่เกิน 50MB)"}

    # ==========================================
    # ตรวจก่อนว่าไฟล์ที่โยนเข้ามาเป็น "ไฟล์ threshold รายเดือน" (มีหลาย sheet
    # ตรงกับ THRESHOLD_SHEETS) หรือเป็นไฟล์ performance ปกติ (1 sheet ต่อไฟล์)
    # ถ้าเป็นไฟล์ threshold -> ประมวลผลแยกแล้ว return เลย ไม่ต้องไปเดาตาราง performance
    # ==========================================
    #
    # หมายเหตุ perf: โหลด workbook แค่ "ครั้งเดียว" ด้วย read_only=True (ไฟล์นี้มี sheet อื่น
    # ในไฟล์เดียวกันที่หนักมาก เช่น หมื่นแถว -- โหลดซ้ำสองรอบ หรือโหลดแบบไม่ read_only จะกิน RAM
    # เป็น GB และเสี่ยงโดน OOM kill)
    # ==========================================
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
        all_sheet_names = wb.sheetnames
    except Exception:
        wb = None
        all_sheet_names = []

    matched_threshold_sheets = [
        s for s in all_sheet_names
        if s.strip().lower() in {k.lower() for k in THRESHOLD_SHEETS}
    ]

    if wb is not None and matched_threshold_sheets:
        try:
            all_records, sheet_report = [], {}
            for sheet_name, prefix in THRESHOLD_SHEETS.items():
                real_name = next(
                    (s for s in wb.sheetnames if s.strip().lower() == sheet_name.strip().lower()),
                    None
                )
                if not real_name:
                    sheet_report[sheet_name] = "not_found"
                    continue
                ws = wb[real_name]
                recs = parse_threshold_sheet(ws, real_name, prefix, header_row_idx=2)
                if not recs:
                    recs = parse_threshold_sheet(ws, real_name, prefix, header_row_idx=1)
                sheet_report[real_name] = f"{len(recs)} raw rows"
                all_records.extend(recs)
            wb.close()

            if not all_records:
                return {
                    "status": "error",
                    "filename": filename,
                    "sheets": sheet_report,
                    "message": f"เจอ sheet threshold ({', '.join(matched_threshold_sheets)}) แต่ parse แล้วไม่ได้ข้อมูลเลย "
                               f"รายละเอียดต่อ sheet: {sheet_report} — ตรวจสอบว่า header จริงอยู่แถวที่ 2 หรือไม่ "
                               f"และมีคอลัมน์ 'Maximum threshold' / 'Minimum threshold' / 'Measure Object' / 'ME' สะกดตรงหรือเปล่า"
                }

            merged = merge_threshold_records(all_records)
            df_thresh = pd.DataFrame(merged)
            upsert_thresholds_df(df_thresh)

            return {
                "status": "success",
                "filename": filename,
                "detected_table": "thresholds",
                "inserted_rows": len(df_thresh),
                "sheets": sheet_report,
                "message": f"ตรวจพบไฟล์ threshold (sheet: {', '.join(matched_threshold_sheets)}) "
                           f"นำเข้าตาราง 'thresholds' สำเร็จ {len(df_thresh)} รายการ!"
            }
        except Exception as e:
            return {"status": "error", "filename": filename, "message": f"การประมวลผลไฟล์ threshold ล้มเหลว: {str(e)}"}

    # ==========================================
    # ไม่ใช่ไฟล์ threshold -> ไปตาม flow เดิม (ไฟล์ performance รายวัน 1 sheet/ไฟล์)
    # ==========================================
    try:
        df = pd.read_excel(io.BytesIO(contents))
        uploaded_columns = set(df.columns.tolist())

        matched_key = None

        # สเต็ปที่ 1: ตรวจสอบจากชื่อไฟล์ + เช็คโครงสร้างคร่าวๆ
        for key, config in TABLE_REGISTRY.items():
            if config["file_keyword"].lower() in filename.lower():
                if all(col in uploaded_columns for col in config["required_columns"]):
                    matched_key = key
                    break

        # สเต็ปที่ 2: กรณีชื่อไฟล์ถูกเปลี่ยน ค้นหาจากคะแนนลายเซ็นของคอลัมน์แทน
        if not matched_key:
            best_score = 0
            for key, config in TABLE_REGISTRY.items():
                if all(col in uploaded_columns for col in config["required_columns"]):
                    score = len(set(config["required_columns"]).intersection(uploaded_columns))
                    if score > best_score:
                        best_score = score
                        matched_key = key

        if not matched_key:
            return {
                "status": "error",
                "filename": filename,
                "message": "ไม่สามารถระบุประเภทไฟล์ได้ โครงสร้างคอลัมน์ไม่ตรงกับตารางใดๆ ในระบบ",
            }

        # เริ่มกระบวนการ Mapping และทำความสะอาดข้อมูล
        target_config = TABLE_REGISTRY[matched_key]
        target_table = target_config["table_name"]

        # เปลี่ยนชื่อคอลัมน์
        df = df.rename(columns=target_config["mapping"])

        # แปลงข้อมูลเวลา
        for time_col in ['begin_time', 'end_time', 'occurrence_time', 'clear_time']:
            if time_col in df.columns:
                df[time_col] = pd.to_datetime(df[time_col])

        exclude_numeric_cols = [
            'begin_time', 'end_time', 'occurrence_time', 'clear_time',
            'me_name', 'me_ip', 'measure_object', 'alarm_id',
            'alarm_code', 'alarm_code_name', 'alarm_severity'
        ]

        for col in df.columns:
            if col not in exclude_numeric_cols:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        # เพิ่มคอลัมน์ Board, Location ถ้าเป็นไฟล์ประสิทธิภาพ
        if target_config["is_performance"]:
            df = extract_measure_object_details(df)

        # จัดการค่าว่าง
        df = df.where(pd.notnull(df), None)

        # กรองเฉพาะคอลัมน์ที่จะใช้
        allowed_db_columns = list(target_config["mapping"].values())
        if target_config["is_performance"]:
            allowed_db_columns.extend(['board_name', 'location_path', 'sub_function'])

        final_columns = [col for col in allowed_db_columns if col in df.columns]
        df = df[final_columns]

        # อัปโหลดเข้า Database แบบป้องกันข้อมูลซ้ำ
        execute_pg_upsert(
            df=df,
            table_name=target_table,
            conflict_target=target_config["conflict_target"]
        )

        return {
            "status": "success",
            "filename": filename,
            "detected_table": target_table,
            "inserted_rows": len(df),
            "message": f"ระบบนำข้อมูลเข้าสู่ตาราง '{target_table}' สำเร็จเรียบร้อย!"
        }

    except Exception as e:
        return {"status": "error", "filename": filename, "message": f"การประมวลผลล้มเหลว: {str(e)}"}


@app.post("/api/upload")
async def upload_file_automated(
    files: List[UploadFile] = File(...),
    user: dict = Depends(require_roles("owner", "admin")),
):
    """
    รับไฟล์ .xlsx ได้หลายไฟล์พร้อมกันในคำขอเดียว (multi-file upload)
    วนลูปประมวลผลทีละไฟล์ แต่ละไฟล์ระบบจะตรวจชนิดและ route เข้าตารางที่ถูกต้องเองอัตโนมัติ
    (ไม่ขึ้นกับว่าไฟล์ไหนถูกส่งมาก่อน/หลัง หรือมาจากหน้าเว็บไหน)
    เฉพาะสิทธิ์ owner หรือ admin เท่านั้นที่อัปโหลดได้
    """
    results = []
    for f in files:
        contents = await f.read()
        result = _process_single_upload(contents, f.filename)
        results.append(result)

    success_count = sum(1 for r in results if r.get("status") == "success")
    total = len(results)
    overall_status = "success" if success_count == total else ("partial" if success_count else "error")

    return {
        "status": overall_status,
        "total_files": total,
        "success_count": success_count,
        "failed_count": total - success_count,
        "results": results,
        "message": f"อัปโหลดสำเร็จ {success_count}/{total} ไฟล์"
                    + ("" if success_count == total else " (บางไฟล์ล้มเหลว ดูรายละเอียดใน results)"),
    }


@app.post("/api/thresholds/upload")
async def upload_thresholds(
    file: UploadFile = File(...),
    header_row: int = Query(2, description="แถวที่เป็น header จริงในแต่ละ sheet"),
    user: dict = Depends(require_roles("owner", "admin")),
):
    """
    อัปโหลดไฟล์ workbook รายเดือน (ไฟล์เดียวกับที่ใช้ upload ข้อมูล performance ก็ได้ ถ้ามี
    sheet: Control Ratio, Fan Ratio, MSU, Line board, Client board อยู่ในไฟล์เดียวกัน)
    ระบบจะอ่านเฉพาะ 5 sheet ที่อยู่ใน THRESHOLD_SHEETS แล้ว upsert เข้าตาราง thresholds
    """
    MAX_FILE_SIZE = 50 * 1024 * 1024
    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="ไฟล์มีขนาดใหญ่เกินไป (จำกัดไม่เกิน 50MB)")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True, read_only=True)

        all_records = []
        sheet_report = {}
        for sheet_name, prefix in THRESHOLD_SHEETS.items():
            real_name = sheet_name
            if real_name not in wb.sheetnames:
                close = [s for s in wb.sheetnames if s.strip().lower() == sheet_name.strip().lower()]
                if not close:
                    sheet_report[sheet_name] = "not_found"
                    continue
                real_name = close[0]

            ws = wb[real_name]
            recs = parse_threshold_sheet(ws, real_name, prefix, header_row_idx=header_row)
            sheet_report[real_name] = f"{len(recs)} raw rows"
            all_records.extend(recs)

        if not all_records:
            raise HTTPException(
                status_code=400,
                detail=f"ไม่พบข้อมูล threshold ใน sheet ที่คาดไว้เลย ({sheet_report}) "
                       f"ลองระบุ header_row ให้ตรงกับไฟล์จริง"
            )

        merged = merge_threshold_records(all_records)
        df = pd.DataFrame(merged)
        upsert_thresholds_df(df)

        return {
            "status": "success",
            "filename": file.filename,
            "sheets": sheet_report,
            "threshold_rows_upserted": len(df),
            "message": "นำเข้าค่า threshold สำเร็จเรียบร้อย!"
        }
    except HTTPException as http_ex:
        raise http_ex
    except Exception as e:
        return {"status": "error", "message": f"การประมวลผลล้มเหลว: {str(e)}"}


@app.get("/api/thresholds")
def get_thresholds(
    metric_key: str | None = Query(None, description="เช่น fan_fan_speed, client_output_optical_power"),
    me_ip: str | None = Query(None),
    me_name: str | None = Query(None),
    measure_object: str | None = Query(None, description="รองรับ partial match"),
    source_sheet: str | None = Query(None),
    user: dict = Depends(get_current_user),
):
    """ ดึงค่า threshold ให้ frontend ใช้แทนค่าที่ hardcode ไว้ในโค้ด """
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        clauses, params = [], []
        if metric_key:
            clauses.append("metric_key = %s"); params.append(metric_key)
        if me_ip:
            clauses.append("me_ip = %s"); params.append(me_ip)
        if me_name:
            clauses.append("me_name = %s"); params.append(me_name)
        if measure_object:
            clauses.append("measure_object ILIKE %s"); params.append(f"%{measure_object}%")
        if source_sheet:
            clauses.append("source_sheet = %s"); params.append(source_sheet)

        where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        query = f'SELECT * FROM "thresholds" {where_sql} ORDER BY me_name, measure_object, metric_key;'
        cursor.execute(query, params)
        data = cursor.fetchall()
        cursor.close()
        conn.close()
        return {"status": "success", "total_records": len(data), "data": data}
    except Exception as e:
        return {"status": "error", "message": str(e)}


# ==========================================
# สั่งรัน uvicorn: uvicorn fast:app --reload
# ==========================================
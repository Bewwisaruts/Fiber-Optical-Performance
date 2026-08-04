"""
import_thresholds.py
=====================
อ่านค่า Maximum/Minimum threshold จาก 5 sheet ในไฟล์ Excel รายเดือน แล้ว
upsert เข้าตาราง `thresholds` ใน Postgres (ดู 01_create_thresholds_table.sql)

Sheet เป้าหมาย (แก้ชื่อได้ที่ TARGET_SHEETS ด้านล่างถ้าชื่อจริงสะกดต่างเล็กน้อย):
    - Control Ratio
    - Fan Ratio
    - MSU
    - Line board
    - Client board

โครงสร้างที่คาดไว้ในแต่ละ sheet (จากภาพตัวอย่างที่ผู้ใช้ส่งมา):
    แถว 1  : ป้ายวันที่ (เช่น 7/8/2026) คร่อมกลุ่มคอลัมน์ -- อ่านเป็น "context" เฉยๆ ไม่ได้ใช้ join key
    แถว 2  : header จริง เช่น 'Site Name', 'ME', 'Measure Object',
             แล้วตามด้วยชุดคอลัมน์ที่ *วนซ้ำ* เป็น 3 คอลัมน์:
                 'Maximum threshold', 'Minimum threshold', '<ชื่อ metric(หน่วย)>'
             ชุดนี้ซ้ำได้หลายรอบต่อ 1 วันที่ (เช่น Output Optical Power, Input Optical Power)
             และซ้ำอีกในวันที่ถัดไป (ค่าควรเท่าเดิม เพราะเป็นค่าคงที่ของอุปกรณ์)

สคริปต์นี้:
    1. เดารูปแบบ header อัตโนมัติ (ไม่ hardcode ตำแหน่งคอลัมน์)
    2. ยุบ (dedupe) ค่าที่ซ้ำกันในแต่ละวันที่ -> เหลือ 1 ค่าต่อ metric ต่อ measure_object
    3. ถ้าเจอค่าไม่ตรงกันระหว่างวันที่ (threshold ถูกแก้ระหว่างเดือน) จะ "เตือน" และ
       ใช้ค่าจากคอลัมน์กลุ่มวันที่ล่าสุด (ขวาสุด) เป็นค่าที่ import
    4. Insert/Update (upsert) เข้าตาราง thresholds ผ่าน staging table เหมือนที่ fast.py ทำ

วิธีใช้:
    # ดูก่อนว่า parser อ่านโครงสร้างไฟล์ถูกไหม (ไม่ยิงเข้า DB)
    python import_thresholds.py path/to/workbook.xlsx --dry-run

    # ของจริง: import เข้า DB
    python import_thresholds.py path/to/workbook.xlsx
"""

import argparse
import os
import re
import sys
from datetime import date, datetime
from collections import defaultdict

import pandas as pd
import openpyxl
from sqlalchemy import create_engine, text

# ==========================================================
# 1. ตั้งค่าการเชื่อมต่อ DB (ให้ตรงกับ fast.py)
# ==========================================================
DB_USER = "postgres"
DB_PASSWORD = "bew30012548"
DB_HOST = "localhost"
DB_PORT = "5432"
DB_NAME = "Client-Card"

CONNECTION_STRING = f"postgresql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"

# ==========================================================
# 2. รายชื่อ sheet เป้าหมาย
#    key   = ชื่อ sheet จริงในไฟล์ (แก้ให้ตรงถ้าตัวสะกดจริงต่างจากนี้)
#    value = prefix ที่จะเติมหน้า metric_key เพื่อกันชื่อชนกันข้าม sheet
# ==========================================================
TARGET_SHEETS = {
    "Control Ratio": "cpu",   # sheet นี้คือ threshold ของ CPU (ยืนยันจากผู้ใช้)
    "Fan Ratio": "fan",
    "MSU": "msu",
    "Line board": "line",
    "Client board": "client",
}

# header ที่ถือเป็นคอลัมน์ "ตัวระบุอุปกรณ์" ไม่ใช่ metric
IDENTIFIER_HEADERS = {
    "site name": "site_name",
    "me": "me_name",
    "me ip": "me_ip",
    "measure object": "measure_object",
}

THRESHOLD_MAX_LABELS = {"maximum threshold", "max threshold"}
THRESHOLD_MIN_LABELS = {"minimum threshold", "min threshold"}


def normalize_header(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def metric_key_from_label(label: str) -> str:
    """ 'Output Optical Power (dBm)' -> 'output_optical_power' """
    base = re.sub(r"\(.*?\)", "", label)          # ตัดหน่วยในวงเล็บออก
    base = re.sub(r"[^a-zA-Z0-9]+", "_", base).strip("_").lower()
    return base or "metric"


def extract_unit(label: str):
    m = re.search(r"\(([^)]+)\)", label)
    return m.group(1) if m else None


def extract_measure_object_parts(measure_object: str):
    board_name = location_path = sub_function = None
    if measure_object:
        m1 = re.match(r"^([^\[]+)", measure_object)
        m2 = re.search(r"\[([0-9-]+)\]", measure_object)
        m3 = re.search(r"\]-([^:()]+)", measure_object)
        board_name = m1.group(1).strip() if m1 else None
        location_path = m2.group(1) if m2 else None
        sub_function = m3.group(1).strip() if m3 else None
    return board_name, location_path, sub_function


def find_date_context(ws, header_row_idx, col_idx):
    """ ไล่หา cell แถวเหนือ header (แถว 1) ที่ใกล้ที่สุดทางซ้ายซึ่งมีค่าเป็นวันที่ """
    row_above = header_row_idx - 1
    if row_above < 1:
        return None
    c = col_idx
    while c >= 1:
        val = ws.cell(row=row_above, column=c).value
        if val not in (None, ""):
            if isinstance(val, (datetime, date)):
                return val if isinstance(val, date) else val.date()
            return val
        c -= 1
    return None


def parse_sheet(ws, sheet_name, prefix, header_row_idx=2):
    """
    คืนค่า list ของ dict แต่ละ record คือ 1 (metric, measure_object) 1 ค่า
    (ยังไม่ dedupe ข้ามวันที่ -- ทำใน merge_records)
    """
    max_col = ws.max_column
    max_row = ws.max_row

    headers = [normalize_header(ws.cell(row=header_row_idx, column=c).value)
               for c in range(1, max_col + 1)]

    # 1) หา identifier columns (ก่อนคอลัมน์ threshold แรก)
    id_cols = {}  # col_idx -> field name
    first_threshold_col = None
    for idx, h in enumerate(headers, start=1):
        low = h.lower()
        if low in THRESHOLD_MAX_LABELS:
            first_threshold_col = idx
            break
        if low in IDENTIFIER_HEADERS:
            id_cols[idx] = IDENTIFIER_HEADERS[low]

    if first_threshold_col is None:
        print(f"  [!] sheet '{sheet_name}': ไม่พบคอลัมน์ 'Maximum threshold' เลย ข้าม sheet นี้")
        return []

    # 2) หา metric blocks: (max_col, min_col, value_col, metric_label)
    blocks = []
    c = first_threshold_col
    while c <= max_col:
        h = headers[c - 1].lower()
        if h in THRESHOLD_MAX_LABELS:
            max_c = c
            min_c = c + 1
            val_c = c + 2
            min_h = headers[min_c - 1].lower() if min_c <= max_col else ""
            if min_h not in THRESHOLD_MIN_LABELS:
                print(f"  [!] sheet '{sheet_name}' col {c}: เจอ '{h}' แต่คอลัมน์ถัดไปไม่ใช่ Minimum threshold ('{headers[min_c-1] if min_c<=max_col else ''}') -- ข้าม")
                c += 1
                continue
            metric_label = headers[val_c - 1] if val_c <= max_col else ""
            if not metric_label or metric_label.lower() in THRESHOLD_MAX_LABELS:
                # ไม่มีคอลัมน์ค่าจริงตามหลัง (ไม่คาดคิด) -- ยังเก็บ threshold ไว้โดยตั้งชื่อ metric แบบ generic
                metric_label = f"metric_{len(blocks)+1}"
                next_c = min_c + 1
            else:
                next_c = val_c + 1
            date_ctx = find_date_context(ws, header_row_idx, max_c)
            blocks.append({
                "max_col": max_c, "min_col": min_c, "val_col": val_c if metric_label and not metric_label.startswith("metric_") else None,
                "metric_label": metric_label, "date": date_ctx,
            })
            c = next_c
        else:
            c += 1

    if not blocks:
        print(f"  [!] sheet '{sheet_name}': หา metric block ไม่เจอ")
        return []

    print(f"  [ok] sheet '{sheet_name}': พบ {len(blocks)} metric block(s) -> "
          + ", ".join(f"{b['metric_label']}@{b['date']}" for b in blocks))

    # 3) เดินทีละแถวข้อมูล (เริ่มแถวถัดจาก header)
    records = []
    for r in range(header_row_idx + 1, max_row + 1):
        row_vals = {}
        for col_idx, field in id_cols.items():
            row_vals[field] = ws.cell(row=r, column=col_idx).value

        measure_object = row_vals.get("measure_object")
        me_name = row_vals.get("me_name")
        if not measure_object or not me_name:
            continue  # แถวว่าง / แถวสุดท้าย

        for b in blocks:
            max_v = ws.cell(row=r, column=b["max_col"]).value
            min_v = ws.cell(row=r, column=b["min_col"]).value
            if max_v is None and min_v is None:
                continue
            metric_key = f"{prefix}_{metric_key_from_label(b['metric_label'])}"
            records.append({
                "source_sheet": sheet_name,
                "metric_key": metric_key,
                "metric_label": b["metric_label"],
                "unit": extract_unit(b["metric_label"]),
                "site_name": row_vals.get("site_name"),
                "me_name": me_name,
                "me_ip": row_vals.get("me_ip"),
                "measure_object": measure_object,
                "min_threshold": min_v,
                "max_threshold": max_v,
                "source_date": b["date"],
            })

    return records


def merge_records(records):
    """
    ยุบหลายแถว (จากหลายช่วงวันที่) ที่เป็น metric+measure_object เดียวกัน ให้เหลือ 1 ค่า
    - ถ้าค่าตรงกันทุกวันที่ -> ใช้เลย
    - ถ้าไม่ตรงกัน -> เตือน และใช้ค่าจาก source_date ล่าสุด (มากสุด)
    """
    groups = defaultdict(list)
    for rec in records:
        key = (rec["metric_key"], rec["me_name"], rec["measure_object"])
        groups[key].append(rec)

    merged = []
    conflicts = 0
    for key, recs in groups.items():
        distinct = {(r["min_threshold"], r["max_threshold"]) for r in recs}
        if len(distinct) > 1:
            conflicts += 1
            # เรียงตามวันที่ (None ไปท้าย) เอาค่าล่าสุด
            recs_sorted = sorted(
                recs, key=lambda r: (r["source_date"] is None, r["source_date"])
            )
            chosen = recs_sorted[-1]
            print(f"  [conflict] {key}: ค่าต่างกันระหว่างวันที่ {distinct} -> ใช้ค่าล่าสุด "
                  f"({chosen['source_date']}) = min {chosen['min_threshold']}, max {chosen['max_threshold']}")
        else:
            chosen = recs[0]
        board_name, location_path, sub_function = extract_measure_object_parts(chosen["measure_object"])
        chosen = {**chosen, "board_name": board_name, "location_path": location_path, "sub_function": sub_function}
        merged.append(chosen)

    if conflicts:
        print(f"\n  รวมพบ {conflicts} metric/measure_object ที่ threshold เปลี่ยนระหว่างเดือน (ใช้ค่าล่าสุดแล้ว)\n")
    return merged


def upsert_thresholds(df: pd.DataFrame, engine):
    if df.empty:
        print("ไม่มีข้อมูลให้ import")
        return

    cols = ["source_sheet", "metric_key", "metric_label", "unit", "site_name",
            "me_name", "me_ip", "measure_object", "board_name", "location_path",
            "sub_function", "min_threshold", "max_threshold", "source_date"]
    df = df[cols]

    staging = "temp_staging_thresholds"
    columns_str = ", ".join(f'"{c}"' for c in cols)

    with engine.begin() as conn:
        df.to_sql(staging, con=conn, if_exists="replace", index=False, method="multi", chunksize=2000)

        upsert_sql = f"""
            INSERT INTO "thresholds" ({columns_str})
            SELECT {columns_str} FROM "{staging}"
            ON CONFLICT (metric_key, me_name, measure_object)
            DO UPDATE SET
                min_threshold = EXCLUDED.min_threshold,
                max_threshold = EXCLUDED.max_threshold,
                unit          = EXCLUDED.unit,
                metric_label  = EXCLUDED.metric_label,
                source_sheet  = EXCLUDED.source_sheet,
                board_name    = EXCLUDED.board_name,
                location_path = EXCLUDED.location_path,
                sub_function  = EXCLUDED.sub_function,
                source_date   = EXCLUDED.source_date,
                updated_at    = now();
        """
        conn.execute(text(upsert_sql))
        conn.execute(text(f'DROP TABLE IF EXISTS "{staging}";'))

    print(f"[done] upsert {len(df)} threshold rows เข้าตาราง thresholds เรียบร้อย")


def main():
    parser = argparse.ArgumentParser(description="Import threshold sheets into Postgres")
    parser.add_argument(
        "xlsx_path",
        nargs="?",
        default="data/07_3BB DWDM_Network Inspection Performance of July_Updated on 16-July-2026.xlsx",
        help="path ของไฟล์ workbook รายเดือน (ค่าเริ่มต้นคือไฟล์ในโฟลเดอร์ data/ ของโปรเจกต์)",
    )
    parser.add_argument("--dry-run", action="store_true",
                         help="แค่ parse และพิมพ์ผลลัพธ์ ไม่ยิงเข้า DB")
    parser.add_argument("--header-row", type=int, default=2,
                         help="แถวที่เป็น header จริง (ค่าเริ่มต้น = 2 ตามภาพตัวอย่าง)")
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx_path):
        print(f"[error] ไม่พบไฟล์: '{args.xlsx_path}'")
        print("        ตรวจสอบว่ารันคำสั่งจาก path เดียวกับที่เห็นในไฟล์ Explorer "
              "(เช่น รันจากโฟลเดอร์ที่มี data/ อยู่ข้างใน) หรือระบุ path เต็มแทน")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)
    print(f"Sheets ในไฟล์: {wb.sheetnames}\n")

    all_records = []
    for sheet_name, prefix in TARGET_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            close = [s for s in wb.sheetnames if s.strip().lower() == sheet_name.strip().lower()]
            if close:
                sheet_name_real = close[0]
                print(f"[i] ใช้ '{sheet_name_real}' แทน '{sheet_name}' (ตัวสะกดใกล้เคียง)")
            else:
                print(f"[!] ไม่พบ sheet '{sheet_name}' ในไฟล์ -- ข้าม")
                continue
        else:
            sheet_name_real = sheet_name

        ws = wb[sheet_name_real]
        recs = parse_sheet(ws, sheet_name_real, prefix, header_row_idx=args.header_row)
        all_records.extend(recs)

    if not all_records:
        print("\nไม่พบ record ใดเลย -- ตรวจสอบชื่อ sheet / แถว header (--header-row) แล้วลองใหม่")
        sys.exit(1)

    merged = merge_records(all_records)
    df = pd.DataFrame(merged)

    print(f"\nสรุป: อ่านได้ {len(all_records)} แถวดิบ -> ยุบเหลือ {len(df)} metric/measure_object ที่ไม่ซ้ำ")
    print(df.groupby("source_sheet")["metric_key"].value_counts())

    if args.dry_run:
        out_csv = "thresholds_preview.csv"
        df.to_csv(out_csv, index=False)
        print(f"\n[dry-run] ไม่ได้เขียนเข้า DB -- ดูตัวอย่างผลลัพธ์ได้ที่ {out_csv}")
        return

    engine = create_engine(CONNECTION_STRING)
    upsert_thresholds(df, engine)


if __name__ == "__main__":
    main()